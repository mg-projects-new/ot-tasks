"""Ongoing P2 rubric: English-only, multi-sheet (Trading Signal + Asset of the day).

Output: workbook with 2 tabs.
  Tab 1 — "{date} Trading Signal":
    rows: Image Title, TG (max 1024), Button, TW, Poll Option 1, Poll Option 2
  Tab 2 — "{date} Asset of the day":
    rows: Asset of the week (header), Image Title (TG, FB), TG (max 1024),
          Button [platform link], FB post
"""

from __future__ import annotations

import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from excel_io.fill_template import apply_sheet_defaults


# Row order for Trading Signal tab
TS_ROWS: list[tuple[str, str]] = [
    ("image_title", "Image Title"),
    ("tg", "TG (max 1024)"),
    ("button", "Button"),
    ("tw", "TW"),
    ("poll_option_1", "Poll Option 1"),
    ("poll_option_2", "Poll Option 2"),
]

# Row order for Asset of the Day tab
AOTD_ROWS: list[tuple[str, str]] = [
    ("image_title", "Image Title (TG, FB)"),
    ("tg", "TG  (max 1024)"),
    ("button", "Button [platform link]"),
    ("fb_post", "FB post"),
]


def ongoing_p2_json_template() -> dict[str, Any]:
    """Two sub-objects, one per tab."""
    return {
        "trading_signal": {k: "" for k, _ in TS_ROWS},
        "asset_of_the_day": {k: "" for k, _ in AOTD_ROWS},
    }


def ongoing_p2_schema_json_text() -> str:
    return json.dumps(ongoing_p2_json_template(), ensure_ascii=False, indent=2)


def fill_trading_signal_tab(ws, data: dict[str, Any], *, jira_url: str) -> None:
    font = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Jira Task Link →"
    ws["A1"].font = font
    ws["A1"].alignment = wrap
    ws["B1"] = jira_url
    ws["B1"].font = font

    ws["B2"] = "EN"
    ws["C2"] = "Chars"
    ws["B2"].font = bold
    ws["C2"].font = bold

    for i, (key, display) in enumerate(TS_ROWS, start=3):
        ws.cell(i, 1, display)
        ws.cell(i, 1).font = font
        ws.cell(i, 1).alignment = wrap
        val = data.get(key, "")
        ws.cell(i, 2, "" if val is None else str(val))
        ws.cell(i, 2).font = font
        ws.cell(i, 2).alignment = wrap
        # Character count only meaningful for posts with limits (TG, TW)
        if key in ("tg", "tw", "image_title"):
            ws.cell(i, 3, f"=LEN(B{i})")
            ws.cell(i, 3).font = font

    # Instructions row
    notes_row = len(TS_ROWS) + 4  # skip one row gap
    instructions = (
        "1. Translate the Image Title row.\n"
        "2. The entire text of the post should be written in one cell.\n"
        "3. Separate paragraphs by one paragraph break (Alt/Option + Enter) as "
        "you'd do that in a doc.\n"
        "4. SM posts can't have text styling. So please don't apply any."
    )
    ws.cell(notes_row, 1, instructions)
    ws.cell(notes_row, 1).font = font
    ws.cell(notes_row, 1).alignment = wrap
    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 8

    apply_sheet_defaults(ws, freeze_row=2)


def fill_asset_of_the_day_tab(ws, data: dict[str, Any], *, jira_url: str) -> None:
    font = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Jira Task Link →"
    ws["A1"].font = font
    ws["A1"].alignment = wrap
    ws["B1"] = jira_url
    ws["B1"].font = font

    ws["B2"] = "EN"
    ws["C2"] = "Chars"
    ws["B2"].font = bold
    ws["C2"].font = bold

    # Header row reading "Asset of the week"
    ws["B3"] = "Asset of the week"
    ws["B3"].font = bold

    # Then the data rows starting at row 4
    for i, (key, display) in enumerate(AOTD_ROWS, start=4):
        ws.cell(i, 1, display)
        ws.cell(i, 1).font = font
        ws.cell(i, 1).alignment = wrap
        val = data.get(key, "")
        ws.cell(i, 2, "" if val is None else str(val))
        ws.cell(i, 2).font = font
        ws.cell(i, 2).alignment = wrap
        if key in ("tg", "image_title", "fb_post"):
            ws.cell(i, 3, f"=LEN(B{i})")
            ws.cell(i, 3).font = font

    notes_row = 4 + len(AOTD_ROWS) + 1
    instructions = (
        "1. Translate the Image Title row.\n"
        "2. The entire text of the post should be written in one cell.\n"
        "3. Separate paragraphs by one paragraph break (Alt/Option + Enter) as "
        "you'd do that in a doc.\n"
        "4. SM posts can't have text styling. So please don't apply any."
    )
    ws.cell(notes_row, 1, instructions)
    ws.cell(notes_row, 1).font = font
    ws.cell(notes_row, 1).alignment = wrap
    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 8

    apply_sheet_defaults(ws, freeze_row=2)


def build_ongoing_p2_workbook(
    data: dict[str, Any], *, jira_url: str,
    ts_date: str = "", aotd_date: str = "",
) -> Workbook:
    wb = Workbook()
    # Rename default sheet to Trading Signal tab
    ts_name = f"{ts_date} Trading Signal" if ts_date else "Trading Signal"
    aotd_name = f"{aotd_date} Asset of the day" if aotd_date else "Asset of the day"
    ts_name = ts_name[:31]
    aotd_name = aotd_name[:31]

    ts_ws = wb.active
    ts_ws.title = ts_name
    fill_trading_signal_tab(ts_ws, data.get("trading_signal") or {}, jira_url=jira_url)

    aotd_ws = wb.create_sheet(aotd_name)
    fill_asset_of_the_day_tab(aotd_ws, data.get("asset_of_the_day") or {}, jira_url=jira_url)

    return wb


def normalize_ongoing_p2_payload(raw: dict[str, Any]) -> dict[str, Any]:
    out = ongoing_p2_json_template()
    if not isinstance(raw, dict):
        return out
    ts = raw.get("trading_signal") or {}
    if isinstance(ts, dict):
        for key in out["trading_signal"]:
            v = ts.get(key)
            if isinstance(v, (str, int, float)):
                out["trading_signal"][key] = str(v)
            elif isinstance(v, dict):
                out["trading_signal"][key] = str(v.get("en", "") or "")
    aotd = raw.get("asset_of_the_day") or {}
    if isinstance(aotd, dict):
        for key in out["asset_of_the_day"]:
            v = aotd.get(key)
            if isinstance(v, (str, int, float)):
                out["asset_of_the_day"][key] = str(v)
            elif isinstance(v, dict):
                out["asset_of_the_day"][key] = str(v.get("en", "") or "")
    return out
