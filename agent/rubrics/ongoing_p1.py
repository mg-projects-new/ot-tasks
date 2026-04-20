"""Ongoing P1 (Economic Calendar) rubric: English-only, single-row deliverable.

Output: one row 'Image Title (AR, TH, FA)' with 6 events joined by newlines.
Each event formatted: Day***Month***WDAY***Time UTC***EventName***Impact
"""

from __future__ import annotations

import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from excel_io.fill_template import apply_sheet_defaults


# Single-row schema
ROW_ORDER: list[tuple[str, str]] = [
    ("image_title", "Image Title  (AR, TH, FA)"),
]


def ongoing_p1_json_template() -> dict[str, Any]:
    """Flat English-only schema: one string for the formatted event list."""
    return {"image_title": ""}


def ongoing_p1_schema_json_text() -> str:
    return json.dumps(ongoing_p1_json_template(), ensure_ascii=False, indent=2)


def fill_ongoing_p1_workbook(wb: Workbook, data: dict[str, Any], *, jira_url: str, week_label: str = "") -> None:
    """Populate workbook: row 1 Jira link, row 2 headers, row 3 image title, row 5 instructions."""
    # Name the sheet like the examples (e.g. "23.03 Economic calendar")
    sheet_name = f"{week_label} Economic calendar" if week_label else "Economic calendar"
    sheet_name = sheet_name[:31]  # Excel sheet name limit
    if wb.sheetnames and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    font = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Row 1: Jira link
    ws["A1"] = "Jira Task Link →"
    ws["A1"].font = font
    ws["A1"].alignment = wrap
    ws["B1"] = jira_url
    ws["B1"].font = font

    # Row 2: headers
    ws["B2"] = "EN"
    ws["C2"] = "Chars"
    ws["B2"].font = bold
    ws["C2"].font = bold
    ws["B2"].alignment = wrap
    ws["C2"].alignment = wrap

    # Row 3: the Image Title row with 6 events
    ws["A3"] = "Image Title  (AR, TH, FA)"
    ws["A3"].font = font
    ws["A3"].alignment = wrap
    val = data.get("image_title", "") or ""
    ws["B3"] = str(val)
    ws["B3"].font = font
    ws["B3"].alignment = wrap
    ws["C3"] = "=LEN(B3)"
    ws["C3"].font = font

    # Row 5: merged instructions row (as in examples)
    instructions = (
        "1. Translate the Image Title row.\n"
        "2. The entire text of the post should be written in one cell.\n"
        "3. Separate paragraphs by one paragraph break (Alt/Option + Enter) as "
        "you'd do that in a doc.\n"
        "4. SM posts can't have text styling. So please don't apply any."
    )
    ws["A5"] = instructions
    ws["A5"].font = font
    ws["A5"].alignment = wrap
    ws.merge_cells("A5:C5")

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100
    ws.column_dimensions["C"].width = 8

    apply_sheet_defaults(ws, freeze_row=2)


def build_ongoing_p1_workbook(data: dict[str, Any], *, jira_url: str, week_label: str = "") -> Workbook:
    wb = Workbook()
    fill_ongoing_p1_workbook(wb, data, jira_url=jira_url, week_label=week_label)
    return wb


def normalize_ongoing_p1_payload(raw: dict[str, Any]) -> dict[str, Any]:
    out = ongoing_p1_json_template()
    if not isinstance(raw, dict):
        return out
    v = raw.get("image_title")
    if isinstance(v, (str, int, float)):
        out["image_title"] = str(v)
    elif isinstance(v, dict):
        out["image_title"] = str(v.get("en", "") or "")
    return out
