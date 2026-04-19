"""Strategy carousel rubric: cards 1–7, Card 8 IG/FB, IG/FB/TG posts."""

from __future__ import annotations

import copy
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from excel_io.fill_template import apply_sheet_defaults

# Locales per deliverable (codes match Excel / brief)
CARD_LOCALES = ["en", "ar", "es", "fr", "hi", "id", "ko", "ms", "pt", "ru", "th", "tr", "vi", "fa"]
IG_POST_LOCALES = ["en", "ar", "th", "fa"]
FB_POST_LOCALES = ["en"]
TG_POST_LOCALES = ["en", "ar", "es", "hi", "id", "ms", "pt", "th", "vi", "fa"]


def _empty_locale_map(locales: list[str]) -> dict[str, str]:
    return {loc: "" for loc in locales}


def strategy_json_template() -> dict[str, Any]:
    """Serializable schema template for prompts (empty strings)."""
    cards = {str(i): _empty_locale_map(CARD_LOCALES) for i in range(1, 8)}
    return {
        "cards": cards,
        "card8_ig": _empty_locale_map(CARD_LOCALES),
        "card8_fb": {"en": ""},
        "ig_post": _empty_locale_map(IG_POST_LOCALES),
        "fb_post": _empty_locale_map(FB_POST_LOCALES),
        "tg_post": _empty_locale_map(TG_POST_LOCALES),
    }


def strategy_schema_json_text() -> str:
    return json.dumps(strategy_json_template(), ensure_ascii=False, indent=2)


def _get_nested(data: dict[str, Any], *keys: str) -> dict[str, Any]:
    cur: Any = data
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return {}
        cur = cur[k]
    return cur if isinstance(cur, dict) else {}


def fill_strategy_workbook(wb: Workbook, data: dict[str, Any], *, jira_url: str) -> None:
    """Populate a Strategy workbook (single sheet 'Лист1')."""
    if "Лист1" in wb.sheetnames:
        ws = wb["Лист1"]
    else:
        ws = wb.active
        ws.title = "Лист1"

    font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    locales = CARD_LOCALES
    ws["A1"] = "Jira Task Link →"
    ws["B1"] = jira_url
    for col, loc in enumerate(locales, start=2):
        ws.cell(2, col, loc)
    ws["A2"] = "Locale →"
    for c in range(1, 2 + len(locales)):
        cell = ws.cell(2, c)
        cell.font = font
        cell.alignment = wrap

    row = 3
    labels_rows: list[tuple[str, dict[str, Any], list[str]]] = []
    for i in range(1, 8):
        card_key = str(i)
        labels_rows.append(
            (f"Card {i}", _get_nested(data, "cards", card_key), CARD_LOCALES),
        )
    labels_rows.append(("Card 8 (IG)", data.get("card8_ig") or {}, CARD_LOCALES))
    labels_rows.append(("Card 8 (FB)", data.get("card8_fb") or {}, FB_POST_LOCALES))
    labels_rows.append(("ig_post", data.get("ig_post") or {}, IG_POST_LOCALES))
    labels_rows.append(("fb_post", data.get("fb_post") or {}, FB_POST_LOCALES))
    labels_rows.append(("tg_post", data.get("tg_post") or {}, TG_POST_LOCALES))

    for label, loc_map, loc_keys in labels_rows:
        ws.cell(row, 1, label)
        for j, loc in enumerate(locales):
            col = j + 2
            val = ""
            if loc in loc_keys:
                raw = loc_map.get(loc)
                val = "" if raw is None else str(raw)
            ws.cell(row, col, val)
            cell = ws.cell(row, col)
            cell.font = font
            cell.alignment = wrap
        ws.cell(row, 1).font = font
        ws.cell(row, 1).alignment = wrap
        row += 1

    apply_sheet_defaults(ws, freeze_row=2)


def build_strategy_workbook(data: dict[str, Any], *, jira_url: str) -> Workbook:
    wb = Workbook()
    fill_strategy_workbook(wb, data, jira_url=jira_url)
    return wb


def normalize_strategy_payload(raw: dict[str, Any]) -> dict[str, Any]:
    """Merge model output with template defaults so missing keys do not break fill."""
    base = strategy_json_template()
    out = copy.deepcopy(base)
    if "cards" in raw and isinstance(raw["cards"], dict):
        for k, v in raw["cards"].items():
            k = str(k)
            if k in out["cards"] and isinstance(v, dict):
                for loc in out["cards"][k]:
                    if loc in v and v[loc] is not None:
                        out["cards"][k][loc] = str(v[loc])
    for key in ("card8_ig", "card8_fb", "ig_post", "fb_post", "tg_post"):
        if key in raw and isinstance(raw[key], dict):
            for loc in out[key]:
                if loc in raw[key] and raw[key][loc] is not None:
                    out[key][loc] = str(raw[key][loc])
    return out
