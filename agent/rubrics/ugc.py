"""UGC (TikTok) rubric: English-only. One workbook per task, multiple sheets (one per video)."""

from __future__ import annotations

import copy
import json
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from excel_io.fill_template import apply_sheet_defaults


# Regex to find "*Видео N.*" or "*Video N.*" blocks in the Jira Description
VIDEO_BLOCK_RE = re.compile(
    r"\*\s*(?:Видео|Video)\s+(\d+)\.\s*\*(.*?)(?=\*\s*(?:Видео|Video)\s+\d+\.\s*\*|\*\s*Дедлайн|\Z)",
    re.DOTALL | re.IGNORECASE,
)

# Inside a video block, find (Scenario NN) or (Script NN) or (Сценарий NN)
SCENARIO_RE = re.compile(
    r"\(\s*(?:Scenario|Script|Сценарий)\s+(\d+)\s*\)",
    re.IGNORECASE,
)


@dataclass
class VideoBlock:
    """Parsed info about one video from the brief."""
    video_number: int                    # 1, 2, 3 ...
    script_id: str = ""                  # Scenario/Script/Сценарий number from brief
    raw_block: str = ""                  # the full text of this video section (for the LLM)


@dataclass
class UGCBriefPlan:
    """What the agent decided to generate from the brief."""
    videos: list[VideoBlock] = field(default_factory=list)
    override_count: int | None = None    # If user passed --videos N


def parse_videos_from_description(description: str) -> list[VideoBlock]:
    """Find all `*Видео N.*` / `*Video N.*` blocks and extract per-video context."""
    if not description:
        return []
    videos: list[VideoBlock] = []
    for match in VIDEO_BLOCK_RE.finditer(description):
        n = int(match.group(1))
        raw = match.group(2).strip()
        scenario = ""
        sm = SCENARIO_RE.search(raw)
        if sm:
            scenario = sm.group(1)
        videos.append(VideoBlock(video_number=n, script_id=scenario, raw_block=raw))
    # Deduplicate by video_number, keep first occurrence, sort
    seen = set()
    unique = []
    for v in videos:
        if v.video_number in seen:
            continue
        seen.add(v.video_number)
        unique.append(v)
    unique.sort(key=lambda v: v.video_number)
    return unique


def plan_from_brief(brief: dict[str, Any], override: int | None = None) -> UGCBriefPlan:
    """Build a UGC plan: parse description, optionally override the count."""
    desc = str(brief.get("Description", "") or "")
    parsed = parse_videos_from_description(desc)
    if override is not None and override > 0:
        # Trim or extend to match override
        if len(parsed) >= override:
            parsed = parsed[:override]
        else:
            # Pad with empty blocks if the user asked for more than the brief has
            for i in range(len(parsed) + 1, override + 1):
                parsed.append(VideoBlock(video_number=i))
    return UGCBriefPlan(videos=parsed, override_count=override)


# ---------- Schema for prompt ----------

def ugc_json_template(plan: UGCBriefPlan) -> dict[str, Any]:
    """Flat, English-only schema for N videos.

    {
      "videos": [
        {"video_number": 1, "script_id": "3", "cover": "", "caption": ""},
        ...
      ]
    }
    """
    videos = [
        {
            "video_number": v.video_number,
            "script_id": v.script_id,
            "cover": "",
            "caption": "",
        }
        for v in plan.videos
    ]
    return {"videos": videos}


def ugc_schema_json_text(plan: UGCBriefPlan) -> str:
    return json.dumps(ugc_json_template(plan), ensure_ascii=False, indent=2)


# ---------- Normalize model output ----------

def normalize_ugc_payload(raw: dict[str, Any], plan: UGCBriefPlan) -> dict[str, Any]:
    """Merge model output with the expected plan; guarantee one entry per planned video."""
    out = ugc_json_template(plan)
    if not isinstance(raw, dict):
        return out
    raw_videos = raw.get("videos")
    if not isinstance(raw_videos, list):
        return out

    # Index by video_number for safe merge
    by_num: dict[int, dict[str, Any]] = {}
    for rv in raw_videos:
        if not isinstance(rv, dict):
            continue
        try:
            n = int(rv.get("video_number"))
        except (TypeError, ValueError):
            continue
        by_num[n] = rv

    for entry in out["videos"]:
        n = entry["video_number"]
        if n not in by_num:
            continue
        rv = by_num[n]
        for key in ("cover", "caption"):
            v = rv.get(key)
            if isinstance(v, str):
                entry[key] = v
            elif v is not None:
                entry[key] = str(v)
        # Respect model's script_id only if brief didn't supply one
        if not entry["script_id"]:
            sid = rv.get("script_id")
            if sid:
                entry["script_id"] = str(sid)

    return out


# ---------- Workbook builder ----------

def _sheet_name(video_number: int, script_id: str) -> str:
    """Match the example naming: 'Video 1 (Script 8)' — or just 'Video 1' if no script id."""
    if script_id:
        return f"Video {video_number} (Script {script_id})"
    return f"Video {video_number}"


def _add_video_sheet(wb: Workbook, *, video_number: int, script_id: str, cover: str, caption: str, jira_url: str) -> None:
    """Create one sheet formatted like the UGC examples. English only (label | EN | Chars)."""
    name = _sheet_name(video_number, script_id)
    # Avoid name collisions (Excel limit: 31 chars, no duplicates)
    base = name[:31]
    candidate = base
    i = 2
    while candidate in wb.sheetnames:
        suffix = f" ({i})"
        candidate = (base[: 31 - len(suffix)] + suffix)
        i += 1
    ws = wb.create_sheet(candidate)

    font = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Row 1: Jira link (merged B1:D1 like the examples)
    ws["A1"] = "Jira Task Link →"
    ws["A1"].font = font
    ws["A1"].alignment = wrap
    ws["B1"] = jira_url
    ws["B1"].font = font
    ws.merge_cells("B1:D1")

    # Row 2: headers
    ws["B2"] = "EN"
    ws["D2"] = "Chars"
    for c in ("B2", "D2"):
        ws[c].font = bold
        ws[c].alignment = wrap

    # Row 3: Cover
    ws["A3"] = "TikTok video - Cover"
    ws["A3"].font = font
    ws["A3"].alignment = wrap
    ws["B3"] = cover or ""
    ws["B3"].font = font
    ws["B3"].alignment = wrap
    ws["D3"] = "=LEN(B3)"
    ws["D3"].font = font
    ws["D3"].alignment = wrap

    # Row 4: Caption
    ws["A4"] = "TikTok video1"
    ws["A4"].font = font
    ws["A4"].alignment = wrap
    ws["B4"] = caption or ""
    ws["B4"].font = font
    ws["B4"].alignment = wrap
    ws["D4"] = "=LEN(B4)"
    ws["D4"].font = font
    ws["D4"].alignment = wrap

    # Row 13: notes (reused from your template)
    ws["A13"] = (
        "1. Don't translate the Image Title row.  "
        "2. The entire text of the post should be written in one cell.  "
        "3. Separate paragraphs by a blank line within the same cell."
    )
    ws["A13"].font = font
    ws["A13"].alignment = wrap
    ws.merge_cells("A13:D13")

    # Column widths (approx match to examples)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8

    apply_sheet_defaults(ws, freeze_row=2)


def build_ugc_workbook(data: dict[str, Any], *, jira_url: str) -> Workbook:
    """Create one xlsx containing N sheets, one per video."""
    wb = Workbook()
    # Remove the default empty sheet
    default = wb.active
    wb.remove(default)

    videos = data.get("videos") or []
    if not videos:
        # Keep at least one placeholder sheet so the file is valid
        _add_video_sheet(wb, video_number=1, script_id="", cover="", caption="", jira_url=jira_url)
        return wb

    for v in videos:
        _add_video_sheet(
            wb,
            video_number=int(v.get("video_number") or 1),
            script_id=str(v.get("script_id") or ""),
            cover=str(v.get("cover") or ""),
            caption=str(v.get("caption") or ""),
            jira_url=jira_url,
        )
    return wb