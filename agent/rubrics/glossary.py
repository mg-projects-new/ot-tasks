"""Glossary rubric: infer columns/rows from example xlsx; fill copy."""

from __future__ import annotations

import copy
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from excel_io.fill_template import apply_sheet_defaults


@dataclass
class GlossaryLayout:
    """Discovered from a template sheet."""

    locale_to_col: dict[str, int]  # e.g. en -> 2
    chars_cols: set[int] = field(default_factory=set)
    row_labels: dict[str, int] = field(default_factory=dict)  # image_title -> row
    sheet_name: str = "Лист1"


_LABEL_ALIASES: dict[str, str] = {
    "image title": "image_title",
    "ig": "ig",
    "fb": "fb",
    "tg": "tg",
    "tg post": "tg",
    "twitter (260)": "twitter",
    "twitter": "twitter",
    "button": "button",
}


def _norm_a_label(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    if s in _LABEL_ALIASES:
        return _LABEL_ALIASES[s]
    if s.startswith("twitter"):
        return "twitter"
    return None


def discover_locale_columns(ws) -> tuple[dict[str, int], set[int]]:
    """
    Find text columns and locale codes from row 2.
    Pattern: ... [locale | 'Chars'] pairs; implicit 'en' if col B precedes first Chars.
    """
    locale_to_col: dict[str, int] = {}
    chars_cols: set[int] = set()
    max_col = ws.max_column or 0
    for c in range(2, max_col + 1):
        v2 = ws.cell(2, c).value
        if v2 is None:
            continue
        if str(v2).strip().lower() == "chars":
            chars_cols.add(c)
            dc = c - 1
            loc_cell = ws.cell(2, dc).value
            if loc_cell is None or str(loc_cell).strip() == "":
                if dc == 2:
                    loc_code = "en"
                else:
                    loc_code = f"col{dc}"
            else:
                loc_code = str(loc_cell).strip().lower()
            locale_to_col[loc_code] = dc

    if not locale_to_col and max_col >= 2:
        # Fallback: row 1 style without Chars markers — treat row 2 as locale names
        for c in range(2, max_col + 1):
            v = ws.cell(2, c).value
            if v and str(v).strip().lower() not in ("chars",):
                locale_to_col[str(v).strip().lower()] = c

    return locale_to_col, chars_cols


def discover_row_labels(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in range(1, min(ws.max_row or 0, 60) + 1):
        key = _norm_a_label(ws.cell(r, 1).value)
        if key and key not in out:
            out[key] = r
    return out


def load_layout_from_workbook(wb: Workbook) -> GlossaryLayout:
    ws = wb.active
    locs, chars = discover_locale_columns(ws)
    rows = discover_row_labels(ws)
    return GlossaryLayout(
        locale_to_col=locs,
        chars_cols=chars,
        row_labels=rows,
        sheet_name=ws.title,
    )


def glossary_json_template(layout: GlossaryLayout) -> dict[str, Any]:
    """Empty nested dict for all logical rows × locales present in template."""
    locs = sorted(layout.locale_to_col.keys(), key=lambda x: (x != "en", x))
    key_order = ["image_title", "ig", "fb", "tg", "twitter", "button"]
    root: dict[str, Any] = {}
    for lk in key_order:
        if lk not in layout.row_labels:
            continue
        root[lk] = {loc: "" for loc in locs}
    return root


def glossary_schema_json_text(layout: GlossaryLayout) -> str:
    return json.dumps(glossary_json_template(layout), ensure_ascii=False, indent=2)


def _col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _max_column_quick(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return int(wb.active.max_column or 0)
    finally:
        wb.close()


def pick_template_path(examples_dir: Path) -> Path | None:
    """Prefer the widest example sheet so locale columns match production exports."""
    if not examples_dir.is_dir():
        return None
    paths = [
        p
        for p in examples_dir.glob("*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    ]
    if not paths:
        return None
    return max(paths, key=_max_column_quick)


def create_minimal_glossary_workbook() -> Workbook:
    """
    When no Examples exist: wide grid similar to real glossary exports.
    Locales: en + 13 others with Chars columns.
    """
    locales = ["en", "ar", "es", "fr", "hi", "id", "ko", "ms", "pt", "ru", "th", "tr", "vi", "fa"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Jira Task Link →"
    ws["B1"] = ""
    col = 2
    ws.cell(2, 1, None)
    for loc in locales:
        ws.cell(2, col, loc.upper() if loc == "en" else loc.upper())
        ws.cell(2, col).font = font
        ws.cell(2, col).alignment = wrap
        col += 1
        ws.cell(2, col, "Chars")
        ws.cell(2, col).font = font
        ws.cell(2, col).alignment = wrap
        col += 1

    labels = ["Image Title", "IG", "FB", "TG Post", "Twitter (260)", "Button"]
    r0 = 3
    for i, lab in enumerate(labels):
        r = r0 + i
        ws.cell(r, 1, lab)
        ws.cell(r, 1).font = font
        ws.cell(r, 1).alignment = wrap
        cc = 2
        for loc in locales:
            ws.cell(r, cc, "")
            ws.cell(r, cc).font = font
            ws.cell(r, cc).alignment = wrap
            cl = _col_letter(cc)
            ws.cell(r, cc + 1, f"=LEN({cl}{r})")
            ws.cell(r, cc + 1).font = font
            ws.cell(r, cc + 1).alignment = wrap
            cc += 2

    apply_sheet_defaults(ws, freeze_row=2)
    return wb


def fill_glossary_workbook(
    wb: Workbook,
    data: dict[str, Any],
    *,
    jira_url: str,
    layout: GlossaryLayout,
) -> None:
    ws = wb[layout.sheet_name] if layout.sheet_name in wb.sheetnames else wb.active
    ws["B1"] = jira_url
    font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    for logical, row_idx in layout.row_labels.items():
        block = data.get(logical)
        if not isinstance(block, dict):
            block = {}
        for loc, col_idx in layout.locale_to_col.items():
            val = block.get(loc, "")
            text = "" if val is None else str(val)
            cell = ws.cell(row_idx, col_idx, text)
            cell.font = font
            cell.alignment = wrap
            if col_idx + 1 in layout.chars_cols:
                cl = _col_letter(col_idx)
                ws.cell(row_idx, col_idx + 1, f"=LEN({cl}{row_idx})")
                ws.cell(row_idx, col_idx + 1).font = font
                ws.cell(row_idx, col_idx + 1).alignment = wrap

    apply_sheet_defaults(ws, freeze_row=2)


def normalize_glossary_payload(raw: dict[str, Any], layout: GlossaryLayout) -> dict[str, Any]:
    tmpl = glossary_json_template(layout)
    out = copy.deepcopy(tmpl)
    for key, locs in out.items():
        if key not in raw or not isinstance(raw[key], dict):
            continue
        for loc in locs:
            if loc in raw[key] and raw[key][loc] is not None:
                out[key][loc] = str(raw[key][loc])
    return out


def get_glossary_layout(examples_dir: Path) -> GlossaryLayout:
    """Infer column/row mapping from the widest example, or from a built-in minimal grid."""
    path = pick_template_path(examples_dir)
    if path is None:
        wb = create_minimal_glossary_workbook()
        layout = load_layout_from_workbook(wb)
        wb.close()
        return layout
    wb = load_workbook(path)
    layout = load_layout_from_workbook(wb)
    wb.close()
    return layout
