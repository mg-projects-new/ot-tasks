"""Read Jira filter export xlsx (row 1 = headers, row 2 = values)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_jira_brief(path: str | Path) -> dict[str, Any]:
    """
    Load row 1 as header keys and row 2 as values into a flat dict.
    Preserves original header strings (e.g. 'Issue key', 'Description').
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(min_row=1, max_row=2, values_only=True)
        r1 = next(rows)
        r2 = next(rows)
        headers = list(r1)
        values = list(r2)
        if len(values) < len(headers):
            values = values + [None] * (len(headers) - len(values))
    finally:
        wb.close()

    out: dict[str, Any] = {}
    for i, h in enumerate(headers):
        if h is None or str(h).strip() == "":
            continue
        key = str(h).strip()
        val = values[i] if i < len(values) else None
        out[key] = val
    return out
