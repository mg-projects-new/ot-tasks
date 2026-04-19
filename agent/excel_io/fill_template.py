"""Shared Excel helpers: styling and serializing workbooks for LLM context."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet


def apply_sheet_defaults(ws: Worksheet, freeze_row: int = 2) -> None:
    """Arial, wrap text for used range, freeze panes below header row."""
    font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    max_r = max(1, ws.max_row or 1)
    max_c = max(1, ws.max_column or 1)
    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
        for cell in row:
            cell.font = font
            cell.alignment = wrap
    ws.freeze_panes = f"A{freeze_row + 1}"


def _col_letter(n: int) -> str:
    """1-based column index to Excel letter(s)."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def serialize_workbook_for_prompt(path: str | Path) -> str:
    """
    Structured text dump of all sheets: names, merged ranges, headers/values,
    and cell comments. Used as few-shot / format reference for Claude.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=False)
    parts: list[str] = []
    parts.append(f"FILE: {path.name}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"\n=== SHEET: {sheet_name} ===")
        parts.append(f"Dimensions: {ws.dimensions}")
        merged = list(ws.merged_cells.ranges)
        if merged:
            parts.append("Merged ranges:")
            for m in merged:
                parts.append(f"  {m}")
        max_row = min(ws.max_row or 0, 200)
        max_col = min(ws.max_column or 0, 80)
        for r in range(1, max_row + 1):
            row_vals: list[str] = []
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                v = cell.value
                if v is None and cell.comment is None:
                    continue
                coord = f"{_col_letter(c)}{r}"
                if cell.comment:
                    row_vals.append(f"{coord}={v!r} [COMMENT: {cell.comment.text!r}]")
                else:
                    row_vals.append(f"{coord}={v!r}")
            if row_vals:
                parts.append(" | ".join(row_vals))
    wb.close()
    return "\n".join(parts)


def slugify_summary(summary: str | None, max_len: int = 80) -> str:
    """Lowercase slug: non-alphanumeric -> underscore; collapse repeats."""
    if not summary:
        return "untitled"
    s = summary.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return (s or "untitled")[:max_len]
